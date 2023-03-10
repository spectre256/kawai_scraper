from bs4 import BeautifulSoup as bs
import requests
import re
import openpyxl
from os.path import exists
import sys

# takes filename and dictionary of prices
def save_spreadsheet(filename, prices):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active

    # adds title row
    title_row = ["URL", "Name", "MSRP", "MAP"]
    for i in range(1, len(title_row) + 1):
        sheet.cell(row = 1, column = i).value = title_row[i - 1]

    # iterates over items and sets the row correctly
    row_num = 2
    for url, item in prices.items():
        row = [url, item["name"], item["msrp"], item["map"]]
        for i in range(1, len(row) + 1):
            sheet.cell(row = row_num, column = i).value = row[i - 1]

        row_num += 1

    # removes extra rows, as this function reformats the whole sheet
    sheet.delete_rows(row_num, sheet.max_row)

    # save sheet
    wb.save(filename)
    # TODO: return list of changed items

# returns list of urls from spreadsheet
# errors if there's no data to get
def get_urls(sheet):
    urls = []
    if sheet.max_row < 2:
        raise Exception("No data found in spreadsheet")

    for row in range(2, sheet.max_row + 1):
        url = sheet.cell(row = row, column = 1).value
        if url != None:
            urls.append(url)

    return urls

# gets prices from existing spreadsheet
def get_prices(sheet):
    prices = {}

    # iterate over all rows that aren't the title row
    for row in range(2, sheet.max_row + 1):
        url = sheet.cell(row = row, column = 1).value
        if not url:
            continue
        name = sheet.cell(row = row, column = 2).value or ""
        msrp = sheet.cell(row = row, column = 3).value or 0.
        map = sheet.cell(row = row, column = 4).value or 0.
        prices[url] = {
            "name": name,
            "msrp": float(msrp),
            "map": float(map),
        }

    return prices

# ignores all non number and "." characters and returns a float
def parse_float(str):
    return float("".join([chr for chr in str if chr.isdigit() or chr == "."]))

# creates listing with checks for missing values
def make_listing(name_result, msrp_result, map_result):
    listing = {
        "name": "",
        "msrp": 0.0,
        "map": 0.0
    }

    if name_result:
        name = name_result.text.strip()
        listing["name"] = name
    if msrp_result:
        listing["msrp"] = parse_float(msrp_result.text)
    if map_result:
        listing["map"] = parse_float(map_result.text)

    return listing

def scrape_urls(urls):
    prices = {}
    # regex to match different types of pages
    product_page = re.compile(r"https://kawaius.com/product/.*")
    store_page = re.compile(r"https://store.kawaius.com/p(roducts)?/.*")

    for url in urls:
        html = requests.get(url).content
        tree = bs(html, "lxml")
        name_result, msrp_result, map_result = None, None, None

        # scrape web page based on url; different pages have different styles
        if re.match(product_page, url):
            name_result = tree.find("h1", class_ = "product-title product_title entry-title")

            # checks for MSRP
            for tag in tree.find_all("strong"):
                if tag.find("span", string = re.compile("MSRP")):
                    msrp_result = tag
                    break

            # checks for MAP under buy online section if there is one
            for tag in tree.find_all("div", class_ = "box has-hover has-hover box-text-bottom"):
                # TODO: add better error checking
                box, color = None, None
                try:
                    box = tag.find("div", class_ = "box-text text-center").find("div")
                    color = box.find("h2", class_ = "thin-font").find("span")
                except:
                    continue

                if color:
                    color = color.text
                    # only get MAP from black model
                    if re.search(r"[bB]lack", color):
                        button = box.find("a").find("span")
                        if button:
                            map_result = button

        elif re.match(store_page, url):
            name_result = tree.find("div", class_ = "titles")
            msrp_result = tree.find("span", class_ = "prce")
            map_result = tree.find("span", class_ = "sale_price")

        prices[url] = make_listing(name_result, msrp_result, map_result)

    return prices


if __name__ == "__main__":
    # default values
    wb = None
    sheet = None
    filename = "data.xlsx"

    # only takes one argument: a filename for an Excel spreadsheet
    if len(sys.argv) == 2:
        arg = sys.argv[1].strip()
        if exists(arg) and re.match(r".*\.xlsx", arg):
            wb = openpyxl.load_workbook(arg)
            sheet = wb.active
            filename = arg
        else:
            raise Exception("Invalid filename passed as argument")
    else:
        raise Exception(f"Expected 1 argument, received {len(sys.argv) - 1}")

    urls = get_urls(sheet)
    old_prices = get_prices(sheet)
    new_prices = scrape_urls(urls)
    changed_prices = {}
    save_spreadsheet(filename, new_prices)

    # compares items in spreadsheet and ones that were just scraped
    for url in urls:
        if old_prices.get(url) != new_prices.get(url):
            changed_prices[url] = new_prices[url]

    # display changed items
    print("{} item(s) updated\n".format(len(changed_prices)))
    for url, data in changed_prices.items():
        name = data["name"]
        if name != "":
            name += ": "
        print("{}{}\nMSRP: {}\nMAP: {}\n".format(name, url, data["msrp"], data["map"]))

